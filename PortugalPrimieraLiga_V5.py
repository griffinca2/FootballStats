import logging
import random
import threading

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from tkinter import *
import easygui
from datetime import datetime as d
import time
import xlwt
from openpyxl import load_workbook
from openpyxl import Workbook

available = []

pageCount = 0
filenameStats = 'PortPrimieraLiga_V3_' + d.now().strftime('%Y_%m_%d_%H_%M%S') + '.xlsx'
statsBook = Workbook()
wsheet = statsBook.create_sheet('Portugal Primiera Liga Refs')
ws = statsBook.active

wsheet.cell(row=1, column=1, value='Portugal Primiera Liga')
ws.cell(row=1, column=1, value='Portugal Primiera Liga')

ws.cell(row=3, column=1).value = 'Club'
ws.cell(row=3, column=2).value = 'Right-footed'
ws.cell(row=3, column=3).value = 'Left-footed'
ws.cell(row=3, column=4).value = 'Header'
ws.cell(row=3, column=5).value = 'Penalty'
ws.cell(row=3, column=6).value = 'Freekick'
ws.cell(row=3, column=7).value = 'Tap-in'
ws.cell(row=3, column=8).value = 'Var.'
ws.cell(row=3, column=9).value = 'Total'


class PortPrimieraLiga:
    def __init__(self):
        # self.statsBook = load_workbook(self.fileName)
        # self.statsBook = Workbook()
        # self.statsBook.create_sheet('La Liga')
        # wsheet = self.statsBook['La Liga']
        self.refNum = 0


    def mainFunction(self):
        user_agents = ['Instagram 10.34.0 Android (18/4.3; 320dpi; 720x1280; Xiaomi; HM 1SW; armani; qcom; en_US)',
                       'Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 74.0.3729.169 Safari / 537.36'
                       'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'
                       'Mozilla/5.0 (Windows NT 5.1; rv:33.0) Gecko/20100101 Firefox/33.0'
                       'Mozilla/5.0 (Windows NT 5.1; rv:7.0.1) Gecko/20100101 Firefox/7.0.1'
                       'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:54.0) Gecko/20100101 Firefox/54.0'
                       'Opera/9.80 (Windows NT 6.1; WOW64) Presto/2.12.388 Version/12.18'
                       'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36 OPR/43.0.2442.991'
                       'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.75 Safari/537.36 OPR/36.0.2130.32'
                       ]
        ua = random.choice(user_agents)
        # self.textfile = open(readfile, "r")
        # self.write_to_file = open(writefile, 'a+')
        # self.handles = self.textfile.read().splitlines()

        # Uncomment below to open program in normal browser
        # browser = webdriver.Chrome(executable_path='chromedriver.exe')

        arg = "user-agent=[" + ua + "]"
        logger.info(arg)

        # Options for using headless browser
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument('--disable-gpu')
        # TODO: Added in V5
        chrome_options.set_capability('unhandledPromptBehavior', 'dismiss')
        # chrome_options.add_argument(
        #    "user-agent=[Instagram 10.34.0 Android (18/4.3; 320dpi; 720x1280; Xiaomi; HM 1SW; armani; qcom; en_US)]")
        chrome_options.add_argument(arg)

        self.browser = webdriver.Chrome(chrome_options=chrome_options,
                                        executable_path=r'C:\webdrivers\chromedriver.exe')

        # self.browser.get("https://www.transfermarkt.com/laliga/torverteilungart/wettbewerb/ES1/plus/1")

        # self.runFirst()

        self.getGoalDistribution()
        # self.getOwnGoals2()
        # self.getCorrectScores()

        # self.getMatchesPlayed()

    def getGoalDistribution(self):
        self.browser.get("https://www.transfermarkt.com/liga-nos/torverteilungart/wettbewerb/PO1/plus/1")
        dataTable = self.browser.find_element_by_css_selector('table.items')
        clubStats = []
        clubs = []

        logger.info('Portugal Primier Liga / Liga Nos Goal Distribution')

        for row in dataTable.find_elements_by_css_selector('tr'):
            for cell in row.find_elements_by_tag_name('td'):
                if cell.text != '':
                    logger.info(cell.text)
                    clubStats.append(cell.text)
            logger.info(clubStats)
            if clubStats:
                clubs.append(clubStats)
            clubStats = []

        logger.info(clubs)
        # TODO: Alphabetization
        clubs.sort()
        char = 'Total'
        # clubs.append(clubs.pop(clubs[0].index('Total')))
        for club1 in clubs:
            if char in club1:
                clubs.append(clubs.pop(clubs.index(club1)))
                #return clubs.index(club1), club1.index(char)
            #raise ValueError("'{char}' is not in list".format(char=char))
        # sorted(clubs, key=lambda x: x[0])
        logger.info('Alphabetized Clubs: ' + str(clubs))

        rw = 4
        for club in clubs:
            logger.info('Club: ' + str(club))
            ws.cell(row=rw, column=1).value = club[0]
            ws.cell(row=rw, column=2).value = club[1]
            ws.cell(row=rw, column=3).value = club[2]
            ws.cell(row=rw, column=4).value = club[3]
            ws.cell(row=rw, column=5).value = club[4]
            ws.cell(row=rw, column=6).value = club[5]
            ws.cell(row=rw, column=7).value = club[6]
            ws.cell(row=rw, column=8).value = club[7]
            ws.cell(row=rw, column=9).value = club[8]
            rw += 1

        # self.textfile.close()
        # self.write_to_file.close()
        # self.browser.quit()
        # statsBook.save(filenameStats)
        self.getOwnGoals2(statsBook)
        # sys.exit()

    def getOwnGoals2(self, wb):
        # Add Additional column for Own Goals
        self.browser.get("https://www.transfermarkt.com/primeira-liga/eigentorstatistik/wettbewerb/PO1")
        dataTable = self.browser.find_element_by_id('yw4')
        ownGoalStats = []
        clubs = []
        hders = []
        ct = 0

        logger.info('Portugal Primier Liga / Liga Nos Own Goals')

        i = 0

        for row in dataTable.find_elements_by_css_selector('tr'):
            i = 0
            ownGoalStats = []

            if ct == 0:
                for th in row.find_elements_by_tag_name('th'):
                    if th.text != '':
                        hders.append(th.text)
                logger.info('Headers: ' + str(hders))
                ct += 1

            for cell in row.find_elements_by_tag_name('td'):
                logger.info('Count: ' + str(i))
                if cell.text != '':
                    logger.info('cell: ' + str(cell.text))
                    # logger.info('cell: ' + str(cell.content))
                    ownGoalStats.append(cell.text)
                else:
                    print('Cell.text is empty')
                i += 1
                logger.info('Clubs: ' + str(clubs))
            if ownGoalStats != '':
                logger.info('OGS:  ' + str(ownGoalStats))
                clubs.append(ownGoalStats)
        # sheet.write(rw, 8, club[8])
        logger.info('No more rows. Entering Data into Excel...')

        # wb = load_workbook("FootballStatsTest3.xlsx")
        # wb = load_workbook(filenameStats)
        wss = wb.active
        # ws = sheet
        # wss = wb.create_sheet('La Liga Stats 2')

        # Add headers to excel file

        cc = 10
        # Write the own goals header. only 'Own Goals' header
        wss.cell(row=3, column=cc).value = 'Own Goals'

        # Add all headers to excel file
        # for h in hders:
        #    ws.cell(row=4, column=cc).value = h
        #   cc += 1

        # iterate through each row in excel file
        a = 0
        rowCount = 4
        ctt = 0
        logger.info('Teams: ' + str(clubs))
        first_column = wss['A']
        qColumn = wss['J']
        # for row in ws.rows:

        for x in range(len(first_column)):
            # Iterate through rows in ownGoalStats
            logger.info('x:' + str(x))
            for c in clubs:
                if ctt < 1:
                    ctt += 1
                    # logger.info('c: ' + str(c))
                    # logger.info('ctt: ' + str(ctt))
                elif len(c) > 0:
                    count = 0
                    columnCount = 10

                    logger.info('Column value: ' + str(first_column[x].value) + ' compare to team[4]: ' + str(c[0]))
                    if first_column[x].value is None:
                        logger.info('first_column[x].value is None')
                        pass
                    elif c[0] is None:
                        logger.info('c[1]: is None')
                        pass
                    else:
                        if c[0] in first_column[x].value:
                            # If team is in ownGoalStats, add the OwnGoal team stat to table in Excel
                            logger.info('# Own Goals: ' + str(c[1]))
                            wss.cell(row=x + 1, column=columnCount).value = int(c[1])
                            logger.info('Cell value: ' + str(wss.cell(row=x + 1, column=columnCount).value))

                        # If team is in ownGoalStats, add the (ALL) team stats to table in Excel
                        # for stat in c:
                        # mycell= ws.cell(row=rowCount, column=columnCount)
                        # mycell = stat
                        #    logger.info('stat[count]: ' + str(stat))
                        #    logger.info('Row/Column: ' + str(x) + ' ' + str(columnCount))
                        #    ws.cell(row=x + 1, column=columnCount).value = stat
                        #    logger.info('Cell value: ' + str(ws.cell(row=x + 1, column=columnCount).value))
                        #    count += 1
                        #    columnCount += 1
                        # else:
                        #    logger.info('# Own Goals: ' + str(c[7]) + ' ' + str(0))
                        #    ws.cell(row=x + 1, column=columnCount).value = 0
                        rowCount += 1
                        ctt += 1

            a += 1

        for x in range(3, len(qColumn)):
            if qColumn[x].value is None:
                wss.cell(row=x + 1, column=10).value = int('0')
                logger.info('Cell value: ' + str(wss.cell(row=x + 1, column=10).value))

        logger.info('A: ' + str(a))
        logger.info('Done writing to Excel')
        # fname = 'newFootballStatsTest' + d.now().strftime('%Y_%m_%d_%H_%M%S') + '.xlsx'
        # wb.save(filenameStats)
        self.getCorrectScores(wb)
        # sys.exit()

    def getCorrectScores(self, wb):
        logger.info('Sleeping... 4')
        time.sleep(4)
        self.browser.get(
            "https://docs.google.com/spreadsheets/d/e/2PACX-1vQWGaYgSOvJDVCTVawRDtaQJ3ZhhA8m1XX7pPEv1qHFUD1muYcQXPkBVGpVB1GmdgZ6gwPJt-D2Tw7X/pubhtml?gid=78603189&single=true&widget=false&headers=false&chrome=false")
        logger.info('Sleeping... 8')
        time.sleep(8)
        dataTable = self.browser.find_element_by_css_selector('table.waffle.no-grid')
        print('Success')
        print(dataTable)

        headers = []
        teamStats = []
        teams = []
        logger.info('Sleeping... 5')
        time.sleep(5)
        trHold = 0
        testwait = 0
        for tr in dataTable.find_elements_by_css_selector('tr'):
            i = 0
            teamStats = []
            logger.info('tr: ' + str(tr))
            for td in tr.find_elements_by_tag_name('td'):
                # logger.info('Sleeping... 2')
                # time.sleep(1)
                # logger.info('td: ' + str(td.text))
                # logger.info('i: ' + str(i))
                # if i == 0:
                #   pass
                if td.text != '':
                    if trHold < 2:
                        logger.info('trHold < 2')
                        pass
                    if trHold == 2:
                        headers.append(td.text)
                        logger.info('headers: ' + str(headers))
                    else:
                        teamStats.append(td.text)
                        logger.info('teamStats: ' + str(teamStats))
                else:
                    'td is null'
                i += 1
            trHold += 1
            teams.append(teamStats)
            logger.info('teams: ' + str(teams))
        logger.info('No more rows. Entering Data into Excel...')

        # wb = load_workbook("FootballStatsTest3.xlsx")
        # wb = load_workbook(filenameStats)
        wss = wb.active
        # ws = sheet
        # wss = wb.create_sheet('La Liga Stats 2')

        # Add headers to excel file
        cc = 13
        ct = 0
        for h in headers:
            wss.cell(row=3, column=cc).value = h
            cc += 1
            ct += 1

        # iterate through each row in excel file
        a = 0
        rowCount = 4
        ctt = 0
        logger.info('Teams: ' + str(teams))
        first_column = wss['A']
        # for row in ws.rows:

        for x in range(len(first_column)):
            # Iterate through rows in ownGoalStats
            logger.info('x:' + str(x))
            for c in teams:
                if ctt < 4:
                    ctt += 1
                    # logger.info('c: ' + str(c))
                    # logger.info('ctt: ' + str(ctt))
                elif len(c) > 0:
                    count = 0
                    columnCount = 12
                    # logger.info('Team: ' + str(c))

                    # logger.info('Team[0]: ' + str(c[0]))

                    # 0logger.info('ctt: ' + str(ctt))
                    # Compare team name in excel row to team names in ownGoalStats
                    # logger.info('Column value: ' + str(first_column[x].value))
                    logger.info('Column value: ' + str(first_column[x].value) + ' compare to team[0]: ' + str(c[0]))
                    # if first_column[x].value == c[0]:
                    if first_column[x].value is None:
                        logger.info('None')
                        pass
                    else:
                        if c[0] in first_column[x].value or \
                                (c[0] == 'Guimaraes' and first_column[x].value == 'Vitória Guimarães SC') or \
                                (c[0] == 'Setubal' and first_column[x].value == 'Vitória Setúbal FC') or \
                                (c[0] == 'Ferreira' and first_column[x].value == 'FC Paços de Ferreira') or \
                                (c[0] == 'Maritimo' and first_column[x].value == 'CS Marítimo') or \
                                (c[0] == 'Famalicao' and first_column[x].value == 'FC Famalicão'):

                            # If team is in ownGoalStats, add the team stats to table in Excel
                            for stat in c:
                                logger.info('stat[count]: ' + str(stat))
                                logger.info('Row/Column: ' + str(x) + ' ' + str(columnCount))
                                wss.cell(row=x + 1, column=columnCount).value = stat
                                logger.info('Cell value: ' + str(wss.cell(row=x + 1, column=columnCount).value))
                                count += 1
                                columnCount += 1
                        rowCount += 1
                        ctt += 1

            a += 1

        logger.info('A: ' + str(a))
        # self.getMatchesPlayed(wb)
        # logger.info('Done writing to Excel. Saving...')
        # wb.save(filenameStats)
        self.getMatchesPlayed(wb)
        # sys.exit()

    def getMatchesPlayed(self, wb):
        # Add Additional column for Own Goals
        self.browser.get("https://www.transfermarkt.com/primeira-liga/spieltagtabelle/wettbewerb/PO1")
        # dataTable = self.browser.find_element_by_xpath('//div[contains(string(), "Table Premier League 19/20")]/table')
        dataTable = self.browser.find_element_by_xpath('//div[contains(string(), "Table Liga NOS 19/20")]/table')

        teams = []
        hders = []
        logger.info('Australian A-League Matches Played')
        ct = 0

        for row in dataTable.find_elements_by_css_selector('tr'):
            i = 0
            teamStats = []
            if ct == 0:
                for th in row.find_elements_by_tag_name('th'):
                    if th.text != '':
                        if th.text == ' ':
                            hders.append('Matches')
                        else:
                            hders.append(th.text)
                logger.info('Headers: ' + str(hders))
                ct += 1

            for cell in row.find_elements_by_tag_name('td'):
                logger.info('Count: ' + str(i))
                if cell.text != '':
                    logger.info('cell: ' + str(cell.text.rstrip()))
                    # logger.info('cell: ' + str(cell.content))
                    teamStats.append(cell.text.rstrip())
            if teamStats != '':
                logger.info('OGS:  ' + str(teamStats))
                teams.append(teamStats)
            i += 1
            logger.info('Teams: ' + str(teams))
        logger.info('No more rows. Entering Data into Excel...')

        # wb = load_workbook("FootballStatsTest3.xlsx")
        # wb = load_workbook(filenameStats)
        wss = wb.active
        # ws = sheet
        # wss = wb.create_sheet('La Liga Stats 2')

        # Add headers to excel file
        cc = 42
        ct = 0
        for h in hders:
            wss.cell(row=3, column=cc).value = h
            cc += 1
            ct += 1

        # iterate through each row in excel file
        a = 0
        rowCount = 4
        ctt = 0
        logger.info('Teams: ' + str(teams))
        first_column = wss['A']
        # for row in ws.rows:

        for x in range(len(first_column)):
            # Iterate through rows in ownGoalStats
            logger.info('x:' + str(x))
            for c in teams:
                if ctt < 4:
                    ctt += 1
                    # logger.info('c: ' + str(c))
                    # logger.info('ctt: ' + str(ctt))
                elif len(c) > 0:
                    count = 0
                    columnCount = 42
                    # logger.info('Team: ' + str(c))

                    # logger.info('Team[0]: ' + str(c[0]))

                    # 0logger.info('ctt: ' + str(ctt))
                    # Compare team name in excel row to team names in ownGoalStats
                    # logger.info('Column value: ' + str(first_column[x].value))
                    logger.info('Column value: ' + str(first_column[x].value) + ' compare to team[0]: ' + str(c[1]))
                    # if first_column[x].value == c[0]:
                    if first_column[x].value is None:
                        logger.info('None')
                        pass
                    else:
                        if c[1] in first_column[x].value or \
                                (c[1] == 'Vit. Guimarães' and first_column[x].value == 'Vitória Guimarães SC') or \
                                (c[1] == 'Vitória Setúbal' and first_column[x].value == 'Vitória Setúbal FC') or \
                                (c[1] == 'Paços Ferreira' and first_column[x].value == 'FC Paços de Ferreira') or \
                                (c[1] == 'Famalicão' and first_column[x].value == 'FC Famalicão'):
                            # If team is in ownGoalStats, add the team stats to table in Excel
                            for stat in c:
                                logger.info('stat[count]: ' + str(stat))
                                logger.info('Row/Column: ' + str(x) + ' ' + str(columnCount))
                                wss.cell(row=x + 1, column=columnCount).value = stat
                                logger.info('Cell value: ' + str(wss.cell(row=x + 1, column=columnCount).value))
                                count += 1
                                columnCount += 1
                        rowCount += 1
                        ctt += 1

            a += 1

        logger.info('A: ' + str(a))
        logger.info('Done writing to Excel. Saving...')
        # fname = 'newFootballStatsTest' + d.now().strftime('%Y_%m_%d_%H_%M%S') + '.xlsx'
        wb.save(filenameStats)
        # statsBook.save(filenameStats)

        self.getConcededPenalties(wb)
        # sys.exit()

    def getConcededPenalties(self, wb):
        self.browser.get("https://www.transfermarkt.com/liga-nos/topVerursachteElfmeter/wettbewerb/PO1/plus/1")
        dataTable = self.browser.find_element_by_css_selector('table.items')
        clubs = []
        hders = []
        ct = 0

        logger.info('Portugal Primeera Liga Conceded Penalties Goals')

        for row in dataTable.find_elements_by_css_selector('tr'):
            i = 0
            teamStats = []
            if ct == 0:
                for th in row.find_elements_by_tag_name('th'):
                    if th.text != '':
                        if th.text == ' ':
                            hders.append('Matches')
                        else:
                            hders.append(th.text)
                logger.info('Headers: ' + str(hders))
                ct += 1

            for cell in row.find_elements_by_tag_name('td'):
                logger.info('Count: ' + str(i))
                if cell.text != '':
                    logger.info('cell: ' + str(cell.text.rstrip()))
                    # logger.info('cell: ' + str(cell.content))
                    teamStats.append(cell.text.rstrip())
            if teamStats != '':
                logger.info('OGS:  ' + str(teamStats))
                clubs.append(teamStats)
            i += 1
            logger.info('Teams: ' + str(clubs))
        logger.info('No more rows. Entering Data into Excel...')

        # wb = load_workbook("FootballStatsTest3.xlsx")
        # wb = load_workbook(filenameStats)
        ws1 = wb.active
        # ws = sheet
        # wss = wb.create_sheet('La Liga Stats 2')

        # Add headers to excel file
        cc = 52
        ct = 0
        for h in hders:
            ws1.cell(row=3, column=cc).value = h
            cc += 1
            ct += 1

        # iterate through each row in excel file
        a = 0
        rowCount = 4
        ctt = 0
        logger.info('Teams: ' + str(clubs))
        first_column = ws1['A']
        # for row in ws.rows:

        for x in range(len(first_column)):
            # Iterate through rows in ownGoalStats
            logger.info('x:' + str(x))
            for c in clubs:
                if ctt < 4:
                    ctt += 1
                    # logger.info('c: ' + str(c))
                    # logger.info('ctt: ' + str(ctt))
                elif len(c) > 0:
                    count = 0
                    columnCount = 52
                    # logger.info('Team: ' + str(c))

                    # logger.info('Team[0]: ' + str(c[0]))

                    # 0logger.info('ctt: ' + str(ctt))
                    # Compare team name in excel row to team names in ownGoalStats
                    # logger.info('Column value: ' + str(first_column[x].value))
                    logger.info('Column value: ' + str(first_column[x].value) + ' compare to team[0]: ' + str(c[1]))
                    # if first_column[x].value == c[0]:
                    if first_column[x].value is None:
                        logger.info('None')
                        pass
                    else:
                        if c[1] in first_column[x].value:
                            # If team is in ownGoalStats, add the team stats to table in Excel
                            for stat in c:
                                # mycell= ws.cell(row=rowCount, column=columnCount)
                                # mycell = stat
                                logger.info('stat[count]: ' + str(stat))
                                logger.info('Row/Column: ' + str(x) + ' ' + str(columnCount))
                                ws1.cell(row=x + 1, column=columnCount).value = stat
                                logger.info('Cell value: ' + str(ws1.cell(row=x + 1, column=columnCount).value))
                                count += 1
                                columnCount += 1
                        rowCount += 1
                        ctt += 1

            a += 1

        logger.info('A: ' + str(a))
        logger.info('Done writing to Excel. Saving...')
        # fname = 'newFootballStatsTest' + d.now().strftime('%Y_%m_%d_%H_%M%S') + '.xlsx'
        wb.save(filenameStats)
        # statsBook.save(filenameStats)

        # self.getConcededPenalties(wb)
        #sys.exit()

        self.getRefs("https://www.transfermarkt.com/liga-nos/schiedsrichter/wettbewerb/PO1/saison_id/2019/plus/1", 0)

    def getRefDetails(self, url):
        logger.info('Getting Ref Details:')

        self.browser.get(url)

        # TODO: Edit this when copying code to new league
        childElement1 = self.browser.find_elements_by_xpath('//img[@alt="Liga NOS"]')
        #print(len(childElement1))
        cE = childElement1[1]
        parentElement1 = cE.find_element_by_xpath("..")
        parentElement = parentElement1.find_element_by_xpath("..")
        logger.info(parentElement)
        dataTable = parentElement.find_element_by_css_selector("table")

        name = self.browser.find_element_by_class_name("spielername-profil").text

        refs = []
        hders = []
        ct = 0

        for row in dataTable.find_elements_by_css_selector('tr'):
            i = 0
            logger.info('i: ' + str(i))
            ref = []
            if ct == 0:
                for th in row.find_elements_by_tag_name('th'):
                    if th.text != '':
                        if th.text == ' ':
                            # hders.append('Matches')
                            logger.info('Header = empty')
                        else:
                            hders.append(th.text)
                logger.info('Headers: ' + str(hders))
                ct += 1

            for cell in row.find_elements_by_tag_name('td'):
                logger.info('Count: ' + str(i))
                if cell.text != '':
                    logger.info('cell: ' + str(cell.text.rstrip()))
                    # logger.info('cell: ' + str(cell.content))
                    ref.append(cell.text.rstrip())
                i += 1
            if ref != '':
                logger.info('Ref:  ' + str(ref))
                refs.append(ref)
        logger.info('Refs: ' + str(refs))

        logger.info('No more rows. Entering Data into Excel...')
        count = 0
        rw = 4
        if self.refNum == 0:
            minCol = 15
        else:
            minCol = 15 + self.refNum * 10
        maxCol = minCol + 10

        # Print Ref's name
        wsheet.cell(row=2, column=minCol).value = name
        logger.info('Ref: ' + str(name))
        # Add headers to excel file
        cc = 42
        ct = 0
        for h in hders:
            wsheet.cell(row=3, column=minCol + ct).value = h
            cc += 1
            ct += 1
        for f in range(ct, 9):
            if ct == 5:
                wsheet.cell(row=3, column=minCol + ct).value = 'Yellow Cards'
            elif ct == 6:
                wsheet.cell(row=3, column=minCol + ct).value = '2nd Yellow Cards'
            elif ct == 7:
                wsheet.cell(row=3, column=minCol + ct).value = 'Red Cards'
            elif ct == 8:
                wsheet.cell(row=3, column=minCol + ct).value = 'Penalty Kicks'
            ct += 1

        # iterate through each row in excel file
        a = 0
        rowCount = 4
        ctt = 0
        logger.info('Refs: ' + str(refs))
        first_column = wsheet['A']
        # for row in ws.rows:

        for ref in refs:
            logger.info('refs: ' + str(refs))
            if len(ref) != 0 and len(ref) != 1:
                # if count != 0:

                logger.info('Ref: ' + str(ref))
                track = 0
                for col in range(minCol, maxCol - 1):
                    logger.info('col: ' + str(col))
                    logger.info('track: ' + str(track))
                    logger.info('minCol: ' + str(minCol))
                    logger.info('maxCol: ' + str(maxCol))
                    if ref[track] == '-':
                        logger.info('col: ' + str(col))
                        wsheet.cell(row=rw, column=col).value = 0
                    # TODO: Changed to handle colons
                    else:
                        if ':' in ref[track]:
                            values = ref[track].rsplit(':', 2)
                            total = int(values[0]) + int(values[1])
                            print(values, total)
                            wsheet.cell(row=rw, column=col).value = total
                        else:
                            wsheet.cell(row=rw, column=col).value = ref[track]
                        logger.info('ws.cell(row=' + str(rw) + ', column='+str(col) + '.value = ' + str(ref[track]))
                    track += 1
                rw += 1

        statsBook.save(filenameStats)
        #sys.exit()

    def getRefs(self, url, pgCount):
        # Add Additional column for Own Goals
        self.browser.get(url)
        parent = self.browser.find_element_by_id('yw1')
        dataTable = self.browser.find_element_by_class_name("items")
        refs = []
        hders = []
        logger.info('Portugal Primiera Liga Liga Refs')
        ct = 0
        count = 0

        wsheet.cell(row=3, column=1).value = 'Referee'
        wsheet.cell(row=3, column=2).value = 'Country'
        wsheet.cell(row=3, column=3).value = 'Debut'
        wsheet.cell(row=3, column=4).value = 'Age at Debut'
        wsheet.cell(row=3, column=5).value = 'Appearances'
        wsheet.cell(row=3, column=6).value = 'Yellow Cards'
        wsheet.cell(row=3, column=7).value = 'Yellow Cards Per Game'
        wsheet.cell(row=3, column=8).value = '2nd Yellow Cards'
        wsheet.cell(row=3, column=9).value = '2nd Yellow Cards Per Game'
        wsheet.cell(row=3, column=10).value = 'Red Cards'
        wsheet.cell(row=3, column=11).value = 'Red Cards Per Game'
        wsheet.cell(row=3, column=12).value = 'Penalty Kicks'
        wsheet.cell(row=3, column=13).value = 'Penalty Kicks Per Game'

        # Instead of checking for link furing run, get all links at once and send to getRefDetails() as list
        # Get link to ref stats page
        # refStatsPage = cell.find_element_by_css_selector('td.hauptlink').find_element_by_css_selector("a")
        refStatsPages = dataTable.find_elements_by_css_selector('td.hauptlink')
        links = []
        hold = 0
        for pg in refStatsPages:
            linkTo = pg.text
            link = pg.find_element_by_css_selector("a").get_attribute('href')
            if hold % 2 == 0 or hold == 0:
                logger.info('Link: ' + str(linkTo) + ' ' + str(link))
                links.append(link)
            hold += 1
        logger.info('Links: ' + str(links) + ' Amount: ' + str(len(links)))
        # self.getRefDetails(url, wb)

        for row in dataTable.find_elements_by_css_selector('tr'):
            refStats = []
            count = 0
            if ct == 0:
                for th in row.find_elements_by_tag_name('th'):
                    if th.text != '':
                        if th.text == ' ':
                            logger.info('Header is empty. ')
                        else:
                            hders.append(th.text)
                logger.info('Headers: ' + str(hders))
                ct += 1

            for cell in row.find_elements_by_tag_name('td'):
                logger.info('cell.text.rstrip(): ' + str(cell.text.rstrip()))

                if cell.text != '':
                    logger.info(
                        'In normal info statement - Count: ' + str(count) + ' Appending ' + str(cell.text.rstrip))
                    # logger.info('cell: ' + str(cell.text.rstrip()))
                    refStats.append(cell.text.rstrip())
                    logger.info('RefStats: ' + str(refStats))
                else:
                    if len(refStats) > 1:
                        logger.info('Info Unknown - Count: ' + str(count) + ' Appending ' + str('Unknown'))
                        # logger.info('cell: ' + str(cell.text.rstrip()))
                        refStats.append('Unknown')
                        logger.info('RefStats: ' + str(refStats))
                count += 1
            if len(refStats) > 3:
                refs.append(refStats)
            logger.info('Refs: ' + str(refs))
        logger.info('No more rows. Entering Data into Excel...')

        logger.info('Refs: ' + str(refs))
        if pgCount == 0:
            rw = 4
        else:
            rw = (pgCount * 4) + 24
        count = 0

        for ref in refs:
            if len(ref) != 0 and len(ref) != 1:
                # if count != 0:

                logger.info('Ref: ' + str(ref))

                for col in range(1, 14):
                    if ref[col] == '-':
                        wsheet.cell(row=rw, column=col).value = 0
                    # TODO: Changed to account for commas
                    else:
                        if ',' in ref[col]:
                            value = ref[col].replace(',', '.')
                            wsheet.cell(row=rw, column=col).value = value
                        else:
                            wsheet.cell(row=rw, column=col).value = ref[col]
                rw += 1
                count += 1
        logger.info('Done writing to Excel. Saving...')
        statsBook.save(filenameStats)
        for lnk in links:
            self.getRefDetails(lnk)
            self.refNum += 1

        self.getNextPage(url, pgCount)
        # sys.exit()

    def getNextPage(self, url, pgCount):
        self.browser.get(url)
        while 1:
            time.sleep(3)
            nextPage = ''
            try:
                nextPage = self.browser.find_element_by_class_name("naechste-seite")
                nextPagelink = nextPage.find_element_by_css_selector("a").get_attribute('href')
                logger.info('Found Next Page element.')
            except:
                statsBook.save(filenameStats)
                logger.info('No more pages. Exiting...')
                sys.exit()
            link = nextPage.find_element_by_css_selector("a").get_attribute('href')
            logger.info('Link: ' + str(link))
            pgCount += 1
            self.getRefs(nextPagelink, pgCount)


# Creating and running GUI interface
window = Tk()

window.title("Football Stats Scraper")
window.geometry('350x200')
lbl = Label(window, text="Football Stats Scraper")
lbl.grid(column=0, row=0)

readfile = ''
writefile = ''


def clicked1():
    global readfile
    readfile = easygui.fileopenbox()


def clicked2():
    global writefile
    wf = easygui.enterbox(msg='Enter output file name(not including .txt)', title='', default='', strip=True)
    writefile = wf + '.txt'


def clicked3():
    # if not readfile.strip():
    #   logger.info('Input file is null. Please enter real file.')
    # if not writefile.strip():
    #    logger.info('Output file is null. Please enter correct values.')
    # else:
    thread = threading.Thread(target=s.mainFunction())
    thread.start()
    # s.mainFunction(readfile, writefile)


logger = logging.getLogger('server_logger')
file_name = 'app' + d.now().strftime('%Y_%m_%d_%H_%M%S') + '.log'
logger.setLevel(logging.INFO)
fh = logging.FileHandler(file_name, encoding="utf-8")
fh.setLevel(logging.INFO)

# create console handler with a higher log level
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
# create formatter and add it to the handlers
formatter = logging.Formatter('%(asctime)s %(name)-12s %(levelname)-8s %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
ch.setFormatter(formatter)
fh.setFormatter(formatter)
# add the handlers to logger
logger.addHandler(ch)
logger.addHandler(fh)

s = PortPrimieraLiga()

btn_input_file = Button(window, text="Select input file:  ", command=clicked1, height=2, width=15)
btn_output_file = Button(window, text='Select output file: ', command=clicked2, height=2, width=15)
btn_run = Button(window, text='Run Program', command=clicked3, height=2, width=15)

btn_input_file.grid(column=1, row=0)
btn_output_file.grid(column=1, row=1)
btn_run.grid(column=1, row=2)
window.mainloop()
